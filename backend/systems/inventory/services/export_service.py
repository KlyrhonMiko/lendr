import io
import csv
import json
from collections import defaultdict
from datetime import date, datetime
from typing import List, Any, Optional
from sqlalchemy import and_, func, or_
from sqlalchemy.orm import aliased
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from sqlmodel import Session, select
from fastapi.responses import StreamingResponse

from core.models.audit_log import AuditLog
from systems.inventory.models.borrow_request import BorrowRequest
from systems.inventory.models.borrow_request_event import BorrowRequestEvent
from systems.inventory.models.borrow_request_item import BorrowRequestItem
from systems.inventory.models.borrow_request_unit import BorrowRequestUnit
from systems.inventory.models.borrow_request_batch import BorrowRequestBatch
from systems.inventory.models.inventory import InventoryItem
from systems.inventory.models.inventory_unit import InventoryUnit
from systems.inventory.models.inventory_batch import InventoryBatch
from systems.inventory.models.inventory_movement import InventoryMovement
from systems.inventory.schemas.import_export_schemas import TimelineMode
from utils.time_utils import normalize_time_window

class ExportService:
    _FORMULA_PREFIXES = ("=", "+", "-", "@")
    _CONTROL_PREFIXES = ("\t", "\r", "\n")
    _MAX_QUERY_ROWS = 5000
    _MAX_RECEIPT_RENDER_ROWS = 1000
    _MAX_EXPORT_ITEMS = 2000
    _MAX_EXPORT_ROWS = 20000
    _RECEIPT_EXPORT_KEYS = (
        "request_id",
        "transaction_ref",
        "receipt_number",
        "borrower_name",
        "borrower_user_id",
        "customer_name",
        "location_name",
        "released_at",
        "released_by_name",
        "expected_return_at",
        "is_emergency",
        "approval_channel",
        "notes",
        "items",
    )
    _RECEIPT_EXPORT_ITEM_KEYS = (
        "item_id",
        "name",
        "classification",
        "qty_released",
        "serial_numbers",
    )

    def _sanitize_export_cell(self, value: Any) -> Any:
        if not isinstance(value, str) or not value:
            return value

        if value.startswith(self._FORMULA_PREFIXES):
            return f"'{value}"

        # Neutralize tab/newline/carriage-return prefixed variants that spreadsheet apps may interpret.
        if value.startswith(self._CONTROL_PREFIXES):
            return f"'{value}"

        first_non_whitespace = next((char for char in value if not char.isspace()), "")
        if first_non_whitespace in self._FORMULA_PREFIXES:
            return f"'{value}"

        return value

    def _sanitize_export_rows(self, rows: List[List[Any]]) -> List[List[Any]]:
        return [
            [self._sanitize_export_cell(value) for value in row]
            for row in rows
        ]

    def _apply_visibility_filters(
        self,
        statement: Any,
        model: type[Any],
        include_deleted: bool,
        include_archived: bool,
    ) -> Any:
        if not include_deleted and hasattr(model, "is_deleted"):
            statement = statement.where(model.is_deleted.is_(False))
        if not include_archived and hasattr(model, "is_archived"):
            statement = statement.where(model.is_archived.is_(False))
        return statement

    def _has_filter_value(self, value: Any) -> bool:
        if value is None:
            return False
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return bool(value.strip())
        if isinstance(value, (list, tuple, set, dict)):
            return len(value) > 0
        return True

    def _reject_unsupported_filters(
        self,
        export_name: str,
        filters: dict[str, Any],
    ) -> None:
        unsupported = sorted(
            name for name, value in filters.items() if self._has_filter_value(value)
        )
        if unsupported:
            joined = ", ".join(unsupported)
            raise ValueError(
                f"Unsupported filter(s) for {export_name} export: {joined}."
            )

    def _resolve_time_window(
        self,
        timeline_mode: TimelineMode | None,
        anchor_date: date | None,
        date_from: datetime | None,
        date_to: datetime | None,
    ) -> tuple[datetime | None, datetime | None]:
        window = normalize_time_window(
            timeline_mode=timeline_mode,
            anchor_date=anchor_date,
            date_from=date_from,
            date_to=date_to,
        )
        return window.date_from, window.date_to

    def _apply_datetime_window(
        self,
        statement: Any,
        column: Any,
        date_from: datetime | None,
        date_to: datetime | None,
    ) -> Any:
        if date_from:
            statement = statement.where(column >= date_from)
        if date_to:
            statement = statement.where(column <= date_to)
        return statement

    def _execute_bounded_query(
        self,
        session: Session,
        statement: Any,
        limit: int,
        entity_label: str,
    ) -> list[Any]:
        rows = session.exec(statement.limit(limit + 1)).all()
        if len(rows) > limit:
            raise ValueError(
                f"Export limit exceeded for {entity_label}. Maximum {limit} rows per export."
            )
        return rows

    def _append_inventory_row(self, rows: List[List[Any]], row: List[Any]) -> None:
        if len(rows) >= self._MAX_EXPORT_ROWS:
            raise ValueError(
                f"Export limit exceeded for inventory rows. Maximum {self._MAX_EXPORT_ROWS} rows per export."
            )
        rows.append(row)

    def _resolve_inventory_unit_by_serial(
        self,
        session: Session,
        serial_number: str,
        include_deleted: bool,
        include_archived: bool,
    ) -> InventoryUnit | None:
        statement = select(InventoryUnit).where(InventoryUnit.serial_number == serial_number)
        statement = self._apply_visibility_filters(
            statement,
            InventoryUnit,
            include_deleted,
            include_archived,
        )
        return session.exec(statement).first()

    def _get_item_count_by_request_id(
        self,
        session: Session,
        request_ids: list[Any],
        include_deleted: bool,
        include_archived: bool,
    ) -> dict[Any, int]:
        if not request_ids:
            return {}

        item_count_statement = select(
            BorrowRequestItem.borrow_uuid,
            func.count(BorrowRequestItem.id),
        ).where(BorrowRequestItem.borrow_uuid.in_(request_ids))
        item_count_statement = self._apply_visibility_filters(
            item_count_statement,
            BorrowRequestItem,
            include_deleted,
            include_archived,
        ).group_by(BorrowRequestItem.borrow_uuid)

        return {
            borrow_uuid: count
            for borrow_uuid, count in session.exec(item_count_statement).all()
            if borrow_uuid is not None
        }

    def _get_unambiguous_borrow_request_ids_for_unit(
        self,
        session: Session,
        unit: InventoryUnit,
        include_deleted: bool,
        include_archived: bool,
    ) -> set[str]:
        request_uuid_statement = select(BorrowRequestUnit.borrow_uuid).join(
            BorrowRequest,
            BorrowRequestUnit.borrow_uuid == BorrowRequest.id,
        ).where(
            BorrowRequestUnit.unit_uuid == unit.id,
        )
        request_uuid_statement = self._apply_visibility_filters(
            request_uuid_statement,
            BorrowRequestUnit,
            include_deleted,
            include_archived,
        )
        request_uuid_statement = self._apply_visibility_filters(
            request_uuid_statement,
            BorrowRequest,
            include_deleted,
            include_archived,
        )

        request_uuids = [
            request_uuid
            for request_uuid in self._execute_bounded_query(
                session,
                request_uuid_statement,
                self._MAX_QUERY_ROWS,
                "serial-linked borrow request IDs",
            )
            if request_uuid is not None
        ]
        if not request_uuids:
            return set()

        request_unit_count_statement = (
            select(
                BorrowRequestUnit.borrow_uuid,
                func.count(func.distinct(BorrowRequestUnit.unit_uuid)),
            )
            .join(BorrowRequest, BorrowRequestUnit.borrow_uuid == BorrowRequest.id)
            .where(BorrowRequestUnit.borrow_uuid.in_(request_uuids))
            .group_by(BorrowRequestUnit.borrow_uuid)
        )
        request_unit_count_statement = self._apply_visibility_filters(
            request_unit_count_statement,
            BorrowRequestUnit,
            include_deleted,
            include_archived,
        )
        request_unit_count_statement = self._apply_visibility_filters(
            request_unit_count_statement,
            BorrowRequest,
            include_deleted,
            include_archived,
        )

        unambiguous_request_uuids = [
            borrow_uuid
            for borrow_uuid, unit_count in self._execute_bounded_query(
                session,
                request_unit_count_statement,
                self._MAX_QUERY_ROWS,
                "serial-linked borrow request unit counts",
            )
            if borrow_uuid is not None and unit_count == 1
        ]
        if not unambiguous_request_uuids:
            return set()

        request_id_statement = select(BorrowRequest.request_id).where(
            BorrowRequest.id.in_(unambiguous_request_uuids)
        )
        request_id_statement = self._apply_visibility_filters(
            request_id_statement,
            BorrowRequest,
            include_deleted,
            include_archived,
        )
        return {
            request_id
            for request_id in self._execute_bounded_query(
                session,
                request_id_statement,
                self._MAX_QUERY_ROWS,
                "serial-linked borrow request IDs",
            )
            if request_id is not None
        }

    def _format_timestamp(self, value: datetime | None) -> str:
        if value is None:
            return "N/A"
        return value.strftime("%Y-%m-%d %H:%M:%S")

    def _format_optional_timestamp(self, value: datetime | None) -> str:
        if value is None:
            return ""
        return value.strftime("%Y-%m-%d %H:%M:%S")

    def _format_bool(self, value: bool | None) -> str:
        if value is None:
            return ""
        return "Yes" if value else "No"

    def _build_user_map(self, session: Session, actor_ids: set[Any]) -> dict[Any, str]:
        from systems.admin.models.user import User

        normalized_ids = {actor_id for actor_id in actor_ids if actor_id is not None}
        if not normalized_ids:
            return {}

        users = session.exec(select(User).where(User.id.in_(normalized_ids))).all()
        return {
            user.id: f"{user.first_name} {user.last_name} ({user.employee_id or user.user_id})"
            for user in users
        }

    def _format_actor_from_map(self, user_map: dict[Any, str], actor_id: Any) -> str:
        if actor_id is None:
            return ""
        return user_map.get(actor_id, str(actor_id))

    def _days_between(self, start: datetime | None, end: datetime | None) -> str:
        if start is None or end is None:
            return ""
        return str(max((end - start).days, 0))

    def _build_filter_summary_rows(
        self,
        export_name: str,
        format: str,
        timeline_mode: TimelineMode | None,
        anchor_date: date | None,
        date_from: datetime | None,
        date_to: datetime | None,
        include_deleted: bool,
        include_archived: bool,
        extra_filters: dict[str, Any] | None = None,
    ) -> list[list[Any]]:
        rows: list[list[Any]] = [
            ["Report", export_name],
            ["Generated At", self._format_optional_timestamp(datetime.now())],
            ["Format", format],
            ["Timeline Mode", getattr(timeline_mode, "value", timeline_mode) or ""],
            ["Anchor Date", anchor_date.isoformat() if anchor_date else ""],
            ["Date From", self._format_optional_timestamp(date_from)],
            ["Date To", self._format_optional_timestamp(date_to)],
            ["Include Deleted", self._format_bool(include_deleted)],
            ["Include Archived", self._format_bool(include_archived)],
        ]
        if extra_filters:
            rows.extend([[key, value if value is not None else ""] for key, value in extra_filters.items()])
        return rows

    def _create_multi_sheet_response(
        self,
        sheets: list[tuple[str, list[str], list[list[Any]]]],
        filename_prefix: str,
    ) -> StreamingResponse:
        filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        for sheet_name, headers, rows in sheets:
            ws = wb.create_sheet(title=sheet_name[:31])
            sanitized_headers = [self._sanitize_export_cell(header) for header in headers]
            sanitized_rows = self._sanitize_export_rows(rows)
            for col, header in enumerate(sanitized_headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            for row_idx, row_data in enumerate(sanitized_rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
        )

    def _build_receipt_export_payload(self, receipt_payload: dict[str, Any]) -> dict[str, Any]:
        export_payload: dict[str, Any] = {
            key: receipt_payload.get(key)
            for key in self._RECEIPT_EXPORT_KEYS
            if key in receipt_payload and key != "items"
        }

        raw_items = receipt_payload.get("items")
        export_payload["items"] = [
            {
                item_key: item_payload.get(item_key)
                for item_key in self._RECEIPT_EXPORT_ITEM_KEYS
                if item_key in item_payload
            }
            for item_payload in raw_items
            if isinstance(item_payload, dict)
        ] if isinstance(raw_items, list) else []

        return export_payload

    def _get_receipt_rendered_payload(self, receipt_service: Any, session: Session, request_id: str) -> str:
        try:
            receipt = receipt_service.generate_release_receipt(session, request_id)
        except ValueError:
            return ""

        if hasattr(receipt, "model_dump"):
            receipt_payload = receipt.model_dump(mode="json")
        else:
            receipt_payload = receipt

        if isinstance(receipt_payload, dict):
            return json.dumps(self._build_receipt_export_payload(receipt_payload))
        return json.dumps(receipt_payload)

    def export_audit_logs(
        self, 
        session: Session, 
        format: str, 
        report_version: str = "v1",
        from_date: Optional[datetime] = None, 
        to_date: Optional[datetime] = None,
        timeline_mode: TimelineMode | None = None,
        anchor_date: date | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
        include_deleted: bool = False,
        include_archived: bool = False,
    ) -> StreamingResponse:
        effective_from_date = date_from or from_date
        effective_to_date = date_to or to_date
        normalized_from_date, normalized_to_date = self._resolve_time_window(
            timeline_mode=timeline_mode,
            anchor_date=anchor_date,
            date_from=effective_from_date,
            date_to=effective_to_date,
        )

        if report_version == "v2":
            return self._export_audit_logs_v2(
                session=session,
                format=format,
                timeline_mode=timeline_mode,
                anchor_date=anchor_date,
                date_from=normalized_from_date,
                date_to=normalized_to_date,
                include_deleted=include_deleted,
                include_archived=include_archived,
            )

        statement = select(AuditLog)
        statement = self._apply_visibility_filters(
            statement,
            AuditLog,
            include_deleted,
            include_archived,
        )
        statement = self._apply_datetime_window(
            statement,
            AuditLog.created_at,
            normalized_from_date,
            normalized_to_date,
        )
            
        logs = self._execute_bounded_query(
            session,
            statement,
            self._MAX_QUERY_ROWS,
            "audit logs",
        )
        
        headers = ["ID", "Action", "Entity ID", "Entity Type", "Actor", "Timestamp", "Reason"]
        data = [
            [
                str(log.id),
                log.action,
                log.entity_id,
                log.entity_type,
                str(log.actor_id),
                log.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                log.reason_code or ""
            ]
            for log in logs
        ]
        
        return self._create_response(headers, data, format, "audit_logs")

    def export_inventory(
        self, 
        session: Session, 
        format: str,
        report_version: str = "v1",
        timeline_mode: TimelineMode | None = None,
        anchor_date: date | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
        include_deleted: bool = False,
        include_archived: bool = False,
    ) -> StreamingResponse:
        normalized_from_date, normalized_to_date = self._resolve_time_window(
            timeline_mode=timeline_mode,
            anchor_date=anchor_date,
            date_from=date_from,
            date_to=date_to,
        )

        if report_version == "v2":
            return self._export_inventory_v2(
                session=session,
                format=format,
                timeline_mode=timeline_mode,
                anchor_date=anchor_date,
                date_from=normalized_from_date,
                date_to=normalized_to_date,
                include_deleted=include_deleted,
                include_archived=include_archived,
            )

        headers = ["name", "category", "classification", "item_type", "is_trackable", "description", "condition", "quantity", "serial_number", "expiration_date"]
        
        # We export a row for each unit (if trackable) or each batch (if not)
        data = []
        items_statement = self._apply_visibility_filters(
            select(InventoryItem),
            InventoryItem,
            include_deleted,
            include_archived,
        )
        items_statement = self._apply_datetime_window(
            items_statement,
            InventoryItem.created_at,
            normalized_from_date,
            normalized_to_date,
        )
        items = self._execute_bounded_query(
            session,
            items_statement,
            self._MAX_EXPORT_ITEMS,
            "inventory items",
        )

        trackable_item_ids = [item.id for item in items if item.is_trackable]
        non_trackable_item_ids = [item.id for item in items if not item.is_trackable]

        units_by_item_id: dict[Any, list[InventoryUnit]] = defaultdict(list)
        if trackable_item_ids:
            units_statement = select(InventoryUnit).where(
                InventoryUnit.inventory_uuid.in_(trackable_item_ids)
            )
            units_statement = self._apply_visibility_filters(
                units_statement,
                InventoryUnit,
                include_deleted,
                include_archived,
            )
            units = self._execute_bounded_query(
                session,
                units_statement,
                self._MAX_EXPORT_ROWS,
                "inventory units",
            )
            for unit in units:
                units_by_item_id[unit.inventory_uuid].append(unit)

        batches_by_item_id: dict[Any, list[InventoryBatch]] = defaultdict(list)
        if non_trackable_item_ids:
            batches_statement = select(InventoryBatch).where(
                InventoryBatch.inventory_uuid.in_(non_trackable_item_ids)
            )
            batches_statement = self._apply_visibility_filters(
                batches_statement,
                InventoryBatch,
                include_deleted,
                include_archived,
            )
            batches = self._execute_bounded_query(
                session,
                batches_statement,
                self._MAX_EXPORT_ROWS,
                "inventory batches",
            )
            for batch in batches:
                batches_by_item_id[batch.inventory_uuid].append(batch)

        for item in items:
            if item.is_trackable:
                units = units_by_item_id.get(item.id, [])
                if not units:
                    self._append_inventory_row(data, [
                        item.name,
                        item.category or "",
                        item.classification or "",
                        item.item_type or "",
                        "true",
                        "",
                        "",
                        "0",
                        "",
                        ""
                    ])
                for unit in units:
                    self._append_inventory_row(data, [
                        item.name,
                        item.category or "",
                        item.classification or "",
                        item.item_type or "",
                        "true",
                        unit.description or "",
                        unit.condition,
                        "1",
                        unit.serial_number,
                        unit.expiration_date.isoformat() if unit.expiration_date else ""
                    ])
            else:
                batches = batches_by_item_id.get(item.id, [])
                if not batches:
                    self._append_inventory_row(data, [
                        item.name,
                        item.category or "",
                        item.classification or "",
                        item.item_type or "",
                        "false",
                        "",
                        "",
                        "0",
                        "",
                        ""
                    ])
                for batch in batches:
                    self._append_inventory_row(data, [
                        item.name,
                        item.category or "",
                        item.classification or "",
                        item.item_type or "",
                        "false",
                        batch.description or "",
                        "",
                        str(batch.total_qty),
                        "",
                        batch.expiration_date.isoformat() if batch.expiration_date else ""
                    ])
      
        return self._create_response(headers, data, format, "inventory_export")

    def export_borrow_history(
        self, 
        session: Session, 
        format: str, 
        report_version: str = "v1",
        status: Optional[str] = None,
        item_id: Optional[str] = None,
        borrower_id: Optional[str] = None,
        serial_number: Optional[str] = None,
        include_receipt_rendered: bool = False,
        timeline_mode: TimelineMode | None = None,
        anchor_date: date | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
        include_deleted: bool = False,
        include_archived: bool = False,
    ) -> StreamingResponse:
        normalized_from_date, normalized_to_date = self._resolve_time_window(
            timeline_mode=timeline_mode,
            anchor_date=anchor_date,
            date_from=date_from,
            date_to=date_to,
        )

        if report_version == "v2":
            return self._export_borrow_history_v2(
                session=session,
                format=format,
                status=status,
                item_id=item_id,
                borrower_id=borrower_id,
                serial_number=serial_number,
                timeline_mode=timeline_mode,
                anchor_date=anchor_date,
                date_from=normalized_from_date,
                date_to=normalized_to_date,
                include_deleted=include_deleted,
                include_archived=include_archived,
            )

        from systems.admin.models.user import User
        headers = [
            "Request ID",
            "Borrower Name + Employee ID",
            "Item Name",
            "Serial Number",
            "Who approved the request",
            "Who assigned the unit",
            "Unit CONDITION on release",
            "Who released",
            "When released",
            "When returned",
            "Who received the return",
            "Unit CONDITION on return",
        ]
        receipt_service: Any | None = None
        receipt_payload_cache: dict[str, str] = {}
        if include_receipt_rendered:
            from systems.inventory.services.borrow_request_service import BorrowService

            receipt_service = BorrowService()

        query_row_limit = (
            self._MAX_RECEIPT_RENDER_ROWS
            if include_receipt_rendered
            else self._MAX_QUERY_ROWS
        )

        def get_cached_receipt_payload(request_id_value: str) -> str:
            if receipt_service is None:
                return ""
            if request_id_value not in receipt_payload_cache:
                receipt_payload_cache[request_id_value] = self._get_receipt_rendered_payload(
                    receipt_service,
                    session,
                    request_id_value,
                )
            return receipt_payload_cache[request_id_value]
        if receipt_service:
            headers.append("Receipt Rendered")

        normalized_serial_number = serial_number.strip() if serial_number else None

        borrower_user = aliased(User)
        approved_user = aliased(User)
        assigned_user = aliased(User)
        assignment_released_user = aliased(User)
        request_released_user = aliased(User)
        received_return_user = aliased(User)

        statement = (
            select(
                BorrowRequestUnit,
                BorrowRequest,
                InventoryUnit,
                InventoryItem,
                borrower_user,
                approved_user,
                assigned_user,
                assignment_released_user,
                request_released_user,
                received_return_user,
            )
            .join(BorrowRequest, BorrowRequestUnit.borrow_uuid == BorrowRequest.id)
            .outerjoin(InventoryUnit, BorrowRequestUnit.unit_uuid == InventoryUnit.id)
            .outerjoin(InventoryItem, InventoryUnit.inventory_uuid == InventoryItem.id)
            .outerjoin(borrower_user, BorrowRequest.borrower_uuid == borrower_user.id)
            .outerjoin(approved_user, BorrowRequest.approved_by == approved_user.id)
            .outerjoin(assigned_user, BorrowRequestUnit.assigned_by == assigned_user.id)
            .outerjoin(
                assignment_released_user,
                BorrowRequestUnit.released_by == assignment_released_user.id,
            )
            .outerjoin(
                request_released_user,
                BorrowRequest.released_by == request_released_user.id,
            )
            .outerjoin(received_return_user, BorrowRequest.received_by == received_return_user.id)
        )
        statement = self._apply_visibility_filters(
            statement,
            BorrowRequestUnit,
            include_deleted,
            include_archived,
        )
        statement = self._apply_visibility_filters(
            statement,
            BorrowRequest,
            include_deleted,
            include_archived,
        )
        statement = self._apply_visibility_filters(
            statement,
            InventoryUnit,
            include_deleted,
            include_archived,
        )
        statement = self._apply_visibility_filters(
            statement,
            InventoryItem,
            include_deleted,
            include_archived,
        )
        statement = self._apply_datetime_window(
            statement,
            BorrowRequest.request_date,
            normalized_from_date,
            normalized_to_date,
        )

        if status and status != "all":
            statement = statement.where(BorrowRequest.status == status)

        if item_id:
            statement = statement.where(InventoryItem.item_id == item_id)

        if borrower_id:
            statement = statement.where(borrower_user.user_id == borrower_id)

        if normalized_serial_number:
            statement = statement.where(InventoryUnit.serial_number == normalized_serial_number)

        results = self._execute_bounded_query(
            session,
            statement,
            query_row_limit,
            "borrow history",
        )

        def _format_actor(user_obj: User | None) -> str:
            if user_obj is None:
                return "N/A"
            return f"{user_obj.first_name} {user_obj.last_name} ({user_obj.employee_id or user_obj.user_id})"

        data = [
            [
                req.request_id,
                _format_actor(borrower) if borrower else (req.customer_name or "N/A"),
                inventory_item.name if inventory_item else "Unknown Item",
                unit.serial_number if unit and unit.serial_number else "",
                _format_actor(approved),
                _format_actor(assigned),
                assignment.condition_on_release
                or (unit.condition if unit and unit.condition else "N/A"),
                _format_actor(assignment_released or request_released),
                self._format_timestamp(assignment.released_at or req.released_at),
                self._format_timestamp(assignment.returned_at or req.returned_at),
                _format_actor(received_return),
                assignment.condition_on_return or "N/A",
                *(
                    [get_cached_receipt_payload(req.request_id)]
                    if receipt_service
                    else []
                ),
            ]
            for (
                assignment,
                req,
                unit,
                inventory_item,
                borrower,
                approved,
                assigned,
                assignment_released,
                request_released,
                received_return,
            ) in results
        ]
        
        return self._create_response(headers, data, format, "borrow_history")

    def export_movements(
        self, 
        session: Session, 
        format: str, 
        report_version: str = "v1",
        movement_type: Optional[str] = None,
        item_id: Optional[str] = None,
        serial_number: Optional[str] = None,
        timeline_mode: TimelineMode | None = None,
        anchor_date: date | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
        include_deleted: bool = False,
        include_archived: bool = False,
    ) -> StreamingResponse:
        normalized_from_date, normalized_to_date = self._resolve_time_window(
            timeline_mode=timeline_mode,
            anchor_date=anchor_date,
            date_from=date_from,
            date_to=date_to,
        )

        if report_version == "v2":
            return self._export_movements_v2(
                session=session,
                format=format,
                movement_type=movement_type,
                item_id=item_id,
                serial_number=serial_number,
                timeline_mode=timeline_mode,
                anchor_date=anchor_date,
                date_from=normalized_from_date,
                date_to=normalized_to_date,
                include_deleted=include_deleted,
                include_archived=include_archived,
            )

        headers = [
            "Serial Number",
            "Item Name",
            "Who Borrowed",
            "When It was Borrowed",
            "Condition on Release",
            "When It was Returned",
            "Condition on Return",
            "Request ID (As a Reference)",
        ]
        from systems.admin.models.user import User

        normalized_serial_number = serial_number.strip() if serial_number else None
        normalized_movement_type = (movement_type or "all").strip().lower()
        if normalized_movement_type not in {"all", "out", "in"}:
            raise ValueError("movement_type must be one of: all, out, in")

        borrower_user = aliased(User)
        borrowed_timestamp = func.coalesce(
            BorrowRequestUnit.released_at,
            BorrowRequest.released_at,
        )
        borrowed_window_timestamp = func.coalesce(
            BorrowRequestUnit.released_at,
            BorrowRequest.released_at,
            BorrowRequest.request_date,
        )
        returned_timestamp = func.coalesce(
            BorrowRequestUnit.returned_at,
            BorrowRequest.returned_at,
        )

        statement = (
            select(
                BorrowRequestUnit,
                BorrowRequest,
                InventoryUnit,
                InventoryItem,
                borrower_user,
            )
            .join(BorrowRequest, BorrowRequestUnit.borrow_uuid == BorrowRequest.id)
            .outerjoin(InventoryUnit, BorrowRequestUnit.unit_uuid == InventoryUnit.id)
            .outerjoin(InventoryItem, InventoryUnit.inventory_uuid == InventoryItem.id)
            .outerjoin(borrower_user, BorrowRequest.borrower_uuid == borrower_user.id)
        )
        statement = self._apply_visibility_filters(
            statement,
            BorrowRequestUnit,
            include_deleted,
            include_archived,
        )
        statement = self._apply_visibility_filters(
            statement,
            BorrowRequest,
            include_deleted,
            include_archived,
        )
        statement = self._apply_visibility_filters(
            statement,
            InventoryUnit,
            include_deleted,
            include_archived,
        )
        statement = self._apply_visibility_filters(
            statement,
            InventoryItem,
            include_deleted,
            include_archived,
        )

        if normalized_movement_type == "out":
            statement = statement.where(borrowed_timestamp.is_not(None))
            statement = self._apply_datetime_window(
                statement,
                borrowed_timestamp,
                normalized_from_date,
                normalized_to_date,
            )
        elif normalized_movement_type == "in":
            statement = statement.where(returned_timestamp.is_not(None))
            statement = self._apply_datetime_window(
                statement,
                returned_timestamp,
                normalized_from_date,
                normalized_to_date,
            )
        else:
            if normalized_from_date or normalized_to_date:
                borrowed_window_predicates = []
                returned_window_predicates = []

                if normalized_from_date:
                    borrowed_window_predicates.append(
                        borrowed_window_timestamp >= normalized_from_date
                    )
                    returned_window_predicates.append(
                        returned_timestamp >= normalized_from_date
                    )
                if normalized_to_date:
                    borrowed_window_predicates.append(
                        borrowed_window_timestamp <= normalized_to_date
                    )
                    returned_window_predicates.append(
                        returned_timestamp <= normalized_to_date
                    )

                statement = statement.where(
                    or_(
                        and_(*borrowed_window_predicates),
                        and_(*returned_window_predicates),
                    )
                )

        if item_id:
            statement = statement.where(InventoryItem.item_id == item_id)

        if normalized_serial_number:
            statement = statement.where(InventoryUnit.serial_number == normalized_serial_number)

        results = self._execute_bounded_query(
            session,
            statement,
            self._MAX_QUERY_ROWS,
            "equipment history",
        )

        def _format_actor(user_obj: User | None) -> str:
            if user_obj is None:
                return "N/A"
            return f"{user_obj.first_name} {user_obj.last_name} ({user_obj.employee_id or user_obj.user_id})"

        data = [
            [
                unit.serial_number if unit and unit.serial_number else "",
                item.name if item else "Unknown Item",
                _format_actor(borrower) if borrower else (req.customer_name or "N/A"),
                self._format_timestamp(assignment.released_at or req.released_at or req.request_date),
                assignment.condition_on_release
                or (unit.condition if unit and unit.condition else "N/A"),
                self._format_timestamp(assignment.returned_at or req.returned_at),
                assignment.condition_on_return or "N/A",
                req.request_id,
            ]
            for assignment, req, unit, item, borrower in results
        ]
        
        return self._create_response(headers, data, format, "inventory_movements")

    def _export_audit_logs_v2(
        self,
        session: Session,
        format: str,
        timeline_mode: TimelineMode | None,
        anchor_date: date | None,
        date_from: datetime | None,
        date_to: datetime | None,
        include_deleted: bool,
        include_archived: bool,
    ) -> StreamingResponse:
        statement = select(AuditLog)
        statement = self._apply_visibility_filters(
            statement,
            AuditLog,
            include_deleted,
            include_archived,
        )
        statement = self._apply_datetime_window(statement, AuditLog.created_at, date_from, date_to)
        logs = self._execute_bounded_query(session, statement, self._MAX_QUERY_ROWS, "audit logs")

        user_map = self._build_user_map(session, {log.actor_id for log in logs})
        headers = [
            "ID",
            "Audit ID",
            "Action",
            "Entity ID",
            "Entity Type",
            "Actor UUID",
            "Actor",
            "Timestamp",
            "Reason",
        ]
        rows = [
            [
                str(log.id),
                log.audit_id,
                log.action,
                log.entity_id,
                log.entity_type,
                str(log.actor_id) if log.actor_id else "",
                self._format_actor_from_map(user_map, log.actor_id),
                self._format_optional_timestamp(log.created_at),
                log.reason_code or "",
            ]
            for log in logs
        ]

        if format == "csv":
            return self._create_response(headers, rows, format, "audit_logs_report")

        summary_rows = self._build_filter_summary_rows(
            "Audit Logs",
            format,
            timeline_mode,
            anchor_date,
            date_from,
            date_to,
            include_deleted,
            include_archived,
            {"Total Logs": len(logs)},
        )
        for label, values in (
            ("Action", [log.action for log in logs]),
            ("Entity Type", [log.entity_type for log in logs]),
            ("Reason", [log.reason_code or "None" for log in logs]),
        ):
            counts: dict[str, int] = defaultdict(int)
            for value in values:
                counts[value] += 1
            for value, count in sorted(counts.items()):
                summary_rows.append([f"{label}: {value}", count])

        return self._create_multi_sheet_response(
            [
                ("Summary", ["Metric", "Value"], summary_rows),
                ("Audit Logs", headers, rows),
            ],
            "audit_logs_report",
        )

    def _export_inventory_v2(
        self,
        session: Session,
        format: str,
        timeline_mode: TimelineMode | None,
        anchor_date: date | None,
        date_from: datetime | None,
        date_to: datetime | None,
        include_deleted: bool,
        include_archived: bool,
    ) -> StreamingResponse:
        item_statement = self._apply_visibility_filters(
            select(InventoryItem),
            InventoryItem,
            include_deleted,
            include_archived,
        )
        item_statement = self._apply_datetime_window(
            item_statement,
            InventoryItem.created_at,
            date_from,
            date_to,
        )
        items = self._execute_bounded_query(
            session,
            item_statement,
            self._MAX_EXPORT_ITEMS,
            "inventory items",
        )
        item_ids = [item.id for item in items]

        units_by_item_id: dict[Any, list[InventoryUnit]] = defaultdict(list)
        batches_by_item_id: dict[Any, list[InventoryBatch]] = defaultdict(list)
        if item_ids:
            units_statement = self._apply_visibility_filters(
                select(InventoryUnit).where(InventoryUnit.inventory_uuid.in_(item_ids)),
                InventoryUnit,
                include_deleted,
                include_archived,
            )
            for unit in self._execute_bounded_query(
                session,
                units_statement,
                self._MAX_EXPORT_ROWS,
                "inventory units",
            ):
                units_by_item_id[unit.inventory_uuid].append(unit)

            batches_statement = self._apply_visibility_filters(
                select(InventoryBatch).where(InventoryBatch.inventory_uuid.in_(item_ids)),
                InventoryBatch,
                include_deleted,
                include_archived,
            )
            for batch in self._execute_bounded_query(
                session,
                batches_statement,
                self._MAX_EXPORT_ROWS,
                "inventory batches",
            ):
                batches_by_item_id[batch.inventory_uuid].append(batch)

        headers = [
            "name",
            "category",
            "classification",
            "item_type",
            "is_trackable",
            "description",
            "condition",
            "quantity",
            "serial_number",
            "expiration_date",
        ]
        rows: list[list[Any]] = []
        for item in items:
            if item.is_trackable:
                units = units_by_item_id.get(item.id, [])
                if not units:
                    self._append_inventory_row(rows, [item.name, item.category or "", item.classification or "", item.item_type or "", "true", "", "", "0", "", ""])
                for unit in units:
                    self._append_inventory_row(rows, [
                        item.name,
                        item.category or "",
                        item.classification or "",
                        item.item_type or "",
                        "true",
                        unit.description or "",
                        unit.condition or "",
                        "1",
                        unit.serial_number or "",
                        unit.expiration_date.isoformat() if unit.expiration_date else "",
                    ])
            else:
                batches = batches_by_item_id.get(item.id, [])
                if not batches:
                    self._append_inventory_row(rows, [item.name, item.category or "", item.classification or "", item.item_type or "", "false", "", "", "0", "", ""])
                for batch in batches:
                    self._append_inventory_row(rows, [
                        item.name,
                        item.category or "",
                        item.classification or "",
                        item.item_type or "",
                        "false",
                        batch.description or "",
                        "",
                        str(batch.total_qty),
                        "",
                        batch.expiration_date.isoformat() if batch.expiration_date else "",
                    ])

        if format == "csv":
            return self._create_response(headers, rows, format, "inventory_export_report")

        summary_rows = self._build_filter_summary_rows(
            "Inventory Catalog",
            format,
            timeline_mode,
            anchor_date,
            date_from,
            date_to,
            include_deleted,
            include_archived,
            {
                "Total Items": len(items),
                "Trackable Items": sum(1 for item in items if item.is_trackable),
                "Non-Trackable Items": sum(1 for item in items if not item.is_trackable),
                "Unit Rows": sum(len(units) for units in units_by_item_id.values()),
                "Batch Rows": sum(len(batches) for batches in batches_by_item_id.values()),
            },
        )
        for label, values in (
            ("Category", [item.category or "None" for item in items]),
            ("Classification", [item.classification or "None" for item in items]),
            ("Item Type", [item.item_type or "None" for item in items]),
        ):
            counts: dict[str, int] = defaultdict(int)
            for value in values:
                counts[value] += 1
            for value, count in sorted(counts.items()):
                summary_rows.append([f"{label}: {value}", count])

        return self._create_multi_sheet_response(
            [
                ("Summary", ["Metric", "Value"], summary_rows),
                ("Catalog", headers, rows),
            ],
            "inventory_export_report",
        )

    def _export_borrow_history_v2(
        self,
        session: Session,
        format: str,
        status: Optional[str],
        item_id: Optional[str],
        borrower_id: Optional[str],
        serial_number: Optional[str],
        timeline_mode: TimelineMode | None,
        anchor_date: date | None,
        date_from: datetime | None,
        date_to: datetime | None,
        include_deleted: bool,
        include_archived: bool,
    ) -> StreamingResponse:
        from systems.admin.models.user import User

        borrower_user = aliased(User)
        requested_item = aliased(InventoryItem)
        serial_unit = aliased(InventoryUnit)
        normalized_serial_number = serial_number.strip() if serial_number else None

        statement = select(BorrowRequest.id, BorrowRequest.request_date).outerjoin(
            borrower_user,
            BorrowRequest.borrower_uuid == borrower_user.id,
        )
        statement = self._apply_visibility_filters(
            statement,
            BorrowRequest,
            include_deleted,
            include_archived,
        )
        statement = self._apply_datetime_window(
            statement,
            BorrowRequest.request_date,
            date_from,
            date_to,
        )
        if status and status != "all":
            statement = statement.where(BorrowRequest.status == status)
        if borrower_id:
            statement = statement.where(borrower_user.user_id == borrower_id)
        if item_id:
            statement = statement.join(
                BorrowRequestItem,
                BorrowRequestItem.borrow_uuid == BorrowRequest.id,
            ).join(
                requested_item,
                BorrowRequestItem.item_uuid == requested_item.id,
            )
            statement = self._apply_visibility_filters(
                statement,
                BorrowRequestItem,
                include_deleted,
                include_archived,
            )
            statement = self._apply_visibility_filters(
                statement,
                requested_item,
                include_deleted,
                include_archived,
            )
            statement = statement.where(requested_item.item_id == item_id)
        if normalized_serial_number:
            statement = statement.join(
                BorrowRequestUnit,
                BorrowRequestUnit.borrow_uuid == BorrowRequest.id,
            ).join(
                serial_unit,
                BorrowRequestUnit.unit_uuid == serial_unit.id,
            )
            statement = self._apply_visibility_filters(
                statement,
                BorrowRequestUnit,
                include_deleted,
                include_archived,
            )
            statement = self._apply_visibility_filters(
                statement,
                serial_unit,
                include_deleted,
                include_archived,
            )
            statement = statement.where(serial_unit.serial_number == normalized_serial_number)

        request_keys = self._execute_bounded_query(
            session,
            statement.distinct().order_by(BorrowRequest.request_date.desc()),
            self._MAX_QUERY_ROWS,
            "borrow history requests",
        )
        request_ids = [request_id for request_id, _ in request_keys if request_id is not None]
        requests: list[BorrowRequest] = []
        if request_ids:
            request_statement = self._apply_visibility_filters(
                select(BorrowRequest).where(BorrowRequest.id.in_(request_ids)),
                BorrowRequest,
                include_deleted,
                include_archived,
            )
            request_map = {
                request.id: request
                for request in self._execute_bounded_query(
                    session,
                    request_statement,
                    self._MAX_QUERY_ROWS,
                    "borrow history request records",
                )
            }
            requests = [
                request_map[request_id]
                for request_id in request_ids
                if request_id in request_map
            ]

        request_items_by_request_id: dict[Any, list[BorrowRequestItem]] = defaultdict(list)
        item_by_id: dict[Any, InventoryItem] = {}
        unit_rows_by_request_id: dict[Any, list[tuple[BorrowRequestUnit, InventoryUnit | None, InventoryItem | None]]] = defaultdict(list)
        batch_rows_by_request_id: dict[Any, list[tuple[BorrowRequestBatch, InventoryBatch | None, InventoryItem | None]]] = defaultdict(list)
        events_by_request_id: dict[Any, list[BorrowRequestEvent]] = defaultdict(list)

        if request_ids:
            item_statement = self._apply_visibility_filters(
                select(BorrowRequestItem).where(BorrowRequestItem.borrow_uuid.in_(request_ids)),
                BorrowRequestItem,
                include_deleted,
                include_archived,
            )
            request_items = self._execute_bounded_query(
                session,
                item_statement.order_by(BorrowRequestItem.created_at.asc()),
                self._MAX_EXPORT_ROWS,
                "borrow request items",
            )
            item_uuids = {row.item_uuid for row in request_items if row.item_uuid is not None}
            if item_uuids:
                inventory_statement = self._apply_visibility_filters(
                    select(InventoryItem).where(InventoryItem.id.in_(item_uuids)),
                    InventoryItem,
                    include_deleted,
                    include_archived,
                )
                item_by_id = {
                    item.id: item
                    for item in self._execute_bounded_query(
                        session,
                        inventory_statement,
                        self._MAX_EXPORT_ITEMS,
                        "borrow request item details",
                    )
                }
            for request_item in request_items:
                request_items_by_request_id[request_item.borrow_uuid].append(request_item)

            unit_statement = (
                select(BorrowRequestUnit, InventoryUnit, InventoryItem)
                .outerjoin(InventoryUnit, BorrowRequestUnit.unit_uuid == InventoryUnit.id)
                .outerjoin(InventoryItem, InventoryUnit.inventory_uuid == InventoryItem.id)
                .where(BorrowRequestUnit.borrow_uuid.in_(request_ids))
            )
            unit_statement = self._apply_visibility_filters(unit_statement, BorrowRequestUnit, include_deleted, include_archived)
            unit_statement = self._apply_visibility_filters(unit_statement, InventoryUnit, include_deleted, include_archived)
            unit_statement = self._apply_visibility_filters(unit_statement, InventoryItem, include_deleted, include_archived)
            unit_rows = self._execute_bounded_query(
                session,
                unit_statement,
                self._MAX_EXPORT_ROWS,
                "borrow request unit assignments",
            )
            for assignment, unit, inventory_item in unit_rows:
                unit_rows_by_request_id[assignment.borrow_uuid].append((assignment, unit, inventory_item))

            batch_statement = (
                select(BorrowRequestBatch, InventoryBatch, InventoryItem)
                .outerjoin(InventoryBatch, BorrowRequestBatch.batch_uuid == InventoryBatch.id)
                .outerjoin(InventoryItem, InventoryBatch.inventory_uuid == InventoryItem.id)
                .where(BorrowRequestBatch.borrow_uuid.in_(request_ids))
            )
            batch_statement = self._apply_visibility_filters(batch_statement, BorrowRequestBatch, include_deleted, include_archived)
            batch_statement = self._apply_visibility_filters(batch_statement, InventoryBatch, include_deleted, include_archived)
            batch_statement = self._apply_visibility_filters(batch_statement, InventoryItem, include_deleted, include_archived)
            batch_rows = self._execute_bounded_query(
                session,
                batch_statement,
                self._MAX_EXPORT_ROWS,
                "borrow request batch assignments",
            )
            for assignment, batch, inventory_item in batch_rows:
                batch_rows_by_request_id[assignment.borrow_uuid].append((assignment, batch, inventory_item))

            event_statement = self._apply_visibility_filters(
                select(BorrowRequestEvent).where(BorrowRequestEvent.borrow_uuid.in_(request_ids)),
                BorrowRequestEvent,
                include_deleted,
                include_archived,
            )
            events = self._execute_bounded_query(
                session,
                event_statement.order_by(BorrowRequestEvent.occurred_at.asc()),
                self._MAX_EXPORT_ROWS,
                "borrow request events",
            )
            for event in events:
                events_by_request_id[event.borrow_uuid].append(event)

        actor_ids: set[Any] = set()
        for request in requests:
            actor_ids.update(
                {
                    request.borrower_uuid,
                    request.approved_by,
                    request.released_by,
                    request.returned_by,
                    request.received_by,
                    request.closed_by,
                }
            )
        for rows in unit_rows_by_request_id.values():
            for assignment, _, _ in rows:
                actor_ids.update(
                    {
                        assignment.requested_by,
                        assignment.approved_by,
                        assignment.assigned_by,
                        assignment.released_by,
                        assignment.returned_by,
                    }
                )
        for rows in batch_rows_by_request_id.values():
            for assignment, _, _ in rows:
                actor_ids.add(assignment.assigned_by)
        for events in events_by_request_id.values():
            actor_ids.update(event.actor_id for event in events)
        user_map = self._build_user_map(session, actor_ids)

        request_headers = [
            "Request ID",
            "Transaction Ref",
            "Status",
            "Requested At",
            "Expected Return At",
            "Borrower",
            "Customer Name",
            "Location Name",
            "Request Channel",
            "Approval Channel",
            "Emergency",
            "Approved By",
            "Approved At",
            "Released By",
            "Released At",
            "Returned By",
            "Returned At",
            "Received By",
            "Closed By",
            "Closed At",
            "Close Reason",
            "Returned On Time",
            "Days Borrowed",
            "Item Count",
            "Unit Count",
            "Batch Count",
            "Notes",
            "Compliance Follow-up Required",
            "Compliance Follow-up Notes",
        ]
        request_rows = [
            [
                request.request_id,
                request.transaction_ref,
                request.status,
                self._format_optional_timestamp(request.request_date),
                self._format_optional_timestamp(request.return_at),
                self._format_actor_from_map(user_map, request.borrower_uuid) or (request.customer_name or ""),
                request.customer_name or "",
                request.location_name or "",
                request.request_channel,
                request.approval_channel,
                self._format_bool(request.is_emergency),
                self._format_actor_from_map(user_map, request.approved_by),
                self._format_optional_timestamp(request.approved_at),
                self._format_actor_from_map(user_map, request.released_by),
                self._format_optional_timestamp(request.released_at),
                self._format_actor_from_map(user_map, request.returned_by),
                self._format_optional_timestamp(request.returned_at),
                self._format_actor_from_map(user_map, request.received_by),
                self._format_actor_from_map(user_map, request.closed_by),
                self._format_optional_timestamp(request.closed_at),
                request.close_reason or "",
                self._format_bool(request.returned_on_time),
                self._days_between(request.released_at, request.returned_at),
                len(request_items_by_request_id.get(request.id, [])),
                len(unit_rows_by_request_id.get(request.id, [])),
                len(batch_rows_by_request_id.get(request.id, [])),
                request.notes or "",
                self._format_bool(request.compliance_followup_required),
                request.compliance_followup_notes or "",
            ]
            for request in requests
        ]

        line_headers = [
            "Request ID",
            "Transaction Ref",
            "Status",
            "Requested At",
            "Borrower",
            "Item ID",
            "Item Name",
            "Category",
            "Classification",
            "Item Type",
            "Is Trackable",
            "Qty Requested",
            "Qty Assigned/Released",
            "Unit ID",
            "Serial Number",
            "Batch ID",
            "Assigned By",
            "Assigned At",
            "Released By",
            "Released At",
            "Condition On Release",
            "Returned By",
            "Returned At",
            "Received By",
            "Condition On Return",
            "Return Notes",
            "Request Notes",
        ]
        line_rows: list[list[Any]] = []
        for request in requests:
            request_items = request_items_by_request_id.get(request.id, [])
            request_unit_rows = unit_rows_by_request_id.get(request.id, [])
            request_batch_rows = batch_rows_by_request_id.get(request.id, [])
            for request_item in request_items:
                inventory_item = item_by_id.get(request_item.item_uuid)
                matching_units = [
                    row for row in request_unit_rows
                    if row[1] is not None and row[1].inventory_uuid == request_item.item_uuid
                ]
                matching_batches = [
                    row for row in request_batch_rows
                    if row[1] is not None and row[1].inventory_uuid == request_item.item_uuid
                ]
                if matching_units:
                    for assignment, unit, unit_item in matching_units:
                        resolved_item = unit_item or inventory_item
                        self._append_inventory_row(line_rows, [
                            request.request_id,
                            request.transaction_ref,
                            request.status,
                            self._format_optional_timestamp(request.request_date),
                            self._format_actor_from_map(user_map, request.borrower_uuid) or (request.customer_name or ""),
                            resolved_item.item_id if resolved_item else "",
                            resolved_item.name if resolved_item else "Deleted Inventory Item",
                            resolved_item.category if resolved_item else "",
                            resolved_item.classification if resolved_item else "",
                            resolved_item.item_type if resolved_item else "",
                            "true",
                            request_item.qty_requested,
                            1 if assignment.released_at else 0,
                            unit.unit_id if unit else "",
                            unit.serial_number if unit and unit.serial_number else "",
                            "",
                            self._format_actor_from_map(user_map, assignment.assigned_by),
                            self._format_optional_timestamp(assignment.assigned_at),
                            self._format_actor_from_map(user_map, assignment.released_by or request.released_by),
                            self._format_optional_timestamp(assignment.released_at or request.released_at),
                            assignment.condition_on_release or (unit.condition if unit else "") or "",
                            self._format_actor_from_map(user_map, assignment.returned_by or request.returned_by),
                            self._format_optional_timestamp(assignment.returned_at or request.returned_at),
                            self._format_actor_from_map(user_map, request.received_by),
                            assignment.condition_on_return or "",
                            assignment.return_notes or "",
                            request.notes or "",
                        ])
                elif matching_batches:
                    for assignment, batch, batch_item in matching_batches:
                        resolved_item = batch_item or inventory_item
                        self._append_inventory_row(line_rows, [
                            request.request_id,
                            request.transaction_ref,
                            request.status,
                            self._format_optional_timestamp(request.request_date),
                            self._format_actor_from_map(user_map, request.borrower_uuid) or (request.customer_name or ""),
                            resolved_item.item_id if resolved_item else "",
                            resolved_item.name if resolved_item else "Deleted Inventory Item",
                            resolved_item.category if resolved_item else "",
                            resolved_item.classification if resolved_item else "",
                            resolved_item.item_type if resolved_item else "",
                            "false",
                            request_item.qty_requested,
                            assignment.qty_assigned,
                            "",
                            "",
                            batch.batch_id if batch else "",
                            self._format_actor_from_map(user_map, assignment.assigned_by),
                            self._format_optional_timestamp(assignment.assigned_at),
                            self._format_actor_from_map(user_map, request.released_by),
                            self._format_optional_timestamp(assignment.released_at or request.released_at),
                            "",
                            self._format_actor_from_map(user_map, request.returned_by),
                            self._format_optional_timestamp(assignment.returned_at or request.returned_at),
                            self._format_actor_from_map(user_map, request.received_by),
                            "",
                            "",
                            request.notes or "",
                        ])
                else:
                    self._append_inventory_row(line_rows, [
                        request.request_id,
                        request.transaction_ref,
                        request.status,
                        self._format_optional_timestamp(request.request_date),
                        self._format_actor_from_map(user_map, request.borrower_uuid) or (request.customer_name or ""),
                        inventory_item.item_id if inventory_item else "",
                        inventory_item.name if inventory_item else "Deleted Inventory Item",
                        inventory_item.category if inventory_item else "",
                        inventory_item.classification if inventory_item else "",
                        inventory_item.item_type if inventory_item else "",
                        self._format_bool(inventory_item.is_trackable) if inventory_item else "",
                        request_item.qty_requested,
                        0,
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        request.notes or "",
                    ])

        event_headers = ["Request ID", "Event ID", "Event Type", "Actor", "Occurred At", "Note"]
        event_rows = [
            [
                request.request_id,
                event.event_id,
                event.event_type,
                self._format_actor_from_map(user_map, event.actor_id),
                self._format_optional_timestamp(event.occurred_at),
                event.note or "",
            ]
            for request in requests
            for event in events_by_request_id.get(request.id, [])
        ]

        if format == "csv":
            return self._create_response(line_headers, line_rows, format, "borrow_history_report")

        status_counts: dict[str, int] = defaultdict(int)
        for request in requests:
            status_counts[request.status] += 1
        summary_rows = self._build_filter_summary_rows(
            "Borrow Request History",
            format,
            timeline_mode,
            anchor_date,
            date_from,
            date_to,
            include_deleted,
            include_archived,
            {
                "Status Filter": status or "",
                "Borrower Filter": borrower_id or "",
                "Item Filter": item_id or "",
                "Serial Filter": normalized_serial_number or "",
                "Total Requests": len(requests),
                "Total Requested Quantity": sum(item.qty_requested for rows in request_items_by_request_id.values() for item in rows),
                "Total Line Rows": len(line_rows),
                "Emergency Requests": sum(1 for request in requests if request.is_emergency),
                "Returned On Time": sum(1 for request in requests if request.returned_on_time is True),
                "Returned Late": sum(1 for request in requests if request.returned_on_time is False),
            },
        )
        for status_value, count in sorted(status_counts.items()):
            summary_rows.append([f"Status: {status_value}", count])

        return self._create_multi_sheet_response(
            [
                ("Summary", ["Metric", "Value"], summary_rows),
                ("Requests", request_headers, request_rows),
                ("Line Items", line_headers, line_rows),
                ("Events", event_headers, event_rows),
            ],
            "borrow_history_report",
        )

    def _export_movements_v2(
        self,
        session: Session,
        format: str,
        movement_type: Optional[str],
        item_id: Optional[str],
        serial_number: Optional[str],
        timeline_mode: TimelineMode | None,
        anchor_date: date | None,
        date_from: datetime | None,
        date_to: datetime | None,
        include_deleted: bool,
        include_archived: bool,
    ) -> StreamingResponse:
        from systems.admin.models.user import User

        normalized_serial_number = serial_number.strip() if serial_number else None
        normalized_movement_type = (movement_type or "all").strip().lower()
        if normalized_movement_type not in {"all", "out", "in"}:
            raise ValueError("movement_type must be one of: all, out, in")

        actor_name_expr = func.concat(User.first_name, " ", User.last_name).label("actor_name")
        statement = (
            select(
                InventoryMovement,
                InventoryUnit,
                InventoryItem,
                User.user_id,
                actor_name_expr,
            )
            .join(InventoryUnit, InventoryMovement.unit_uuid == InventoryUnit.id)
            .outerjoin(InventoryItem, InventoryMovement.inventory_uuid == InventoryItem.id)
            .outerjoin(User, InventoryMovement.actor_id == User.id)
        )
        statement = self._apply_visibility_filters(statement, InventoryMovement, include_deleted, include_archived)
        statement = self._apply_visibility_filters(statement, InventoryUnit, include_deleted, include_archived)
        statement = self._apply_visibility_filters(statement, InventoryItem, include_deleted, include_archived)
        statement = self._apply_datetime_window(statement, InventoryMovement.occurred_at, date_from, date_to)

        if normalized_movement_type == "out":
            statement = statement.where(InventoryMovement.movement_type == "borrow_release")
        elif normalized_movement_type == "in":
            statement = statement.where(InventoryMovement.movement_type == "borrow_return")
        if item_id:
            statement = statement.where(InventoryItem.item_id == item_id)
        if normalized_serial_number:
            statement = statement.where(InventoryUnit.serial_number == normalized_serial_number)

        movement_rows = self._execute_bounded_query(
            session,
            statement.order_by(InventoryUnit.serial_number.asc(), InventoryMovement.occurred_at.asc()),
            self._MAX_QUERY_ROWS,
            "equipment history lifecycle events",
        )

        borrow_ref_ids = {
            movement.reference_id
            for movement, _, _, _, _ in movement_rows
            if movement.reference_id
            and movement.movement_type in {"borrow_release", "borrow_return"}
            and movement.reference_type in (None, "borrow_request")
        }
        borrow_context: dict[str, tuple[BorrowRequest, str]] = {}
        if borrow_ref_ids:
            borrower_user = aliased(User)
            borrow_statement = (
                select(BorrowRequest, borrower_user)
                .outerjoin(borrower_user, BorrowRequest.borrower_uuid == borrower_user.id)
                .where(BorrowRequest.request_id.in_(borrow_ref_ids))
            )
            borrow_statement = self._apply_visibility_filters(
                borrow_statement,
                BorrowRequest,
                include_deleted,
                include_archived,
            )
            for request, borrower in self._execute_bounded_query(
                session,
                borrow_statement,
                self._MAX_QUERY_ROWS,
                "equipment history borrow context",
            ):
                borrower_name = (
                    f"{borrower.first_name} {borrower.last_name} ({borrower.employee_id or borrower.user_id})"
                    if borrower
                    else (request.customer_name or "")
                )
                borrow_context[request.request_id] = (request, borrower_name)

        request_ids = [request.id for request, _ in borrow_context.values() if request.id is not None]
        assignment_context: dict[tuple[Any, Any], BorrowRequestUnit] = {}
        if request_ids:
            assignment_statement = self._apply_visibility_filters(
                select(BorrowRequestUnit).where(BorrowRequestUnit.borrow_uuid.in_(request_ids)),
                BorrowRequestUnit,
                include_deleted,
                include_archived,
            )
            assignments = self._execute_bounded_query(
                session,
                assignment_statement,
                self._MAX_EXPORT_ROWS,
                "equipment history request assignments",
            )
            for assignment in assignments:
                assignment_context[(assignment.borrow_uuid, assignment.unit_uuid)] = assignment

        event_headers = [
            "Serial Number",
            "Unit ID",
            "Item ID",
            "Item Name",
            "Movement ID",
            "Event Time",
            "Lifecycle Event Type",
            "Quantity Change",
            "Reason Code",
            "Note",
            "Actor",
            "Reference Type",
            "Reference ID",
            "Request ID",
            "Borrower",
            "Customer Name",
            "Location Name",
            "Released At",
            "Returned At",
            "Condition On Release",
            "Condition On Return",
            "Reversed",
        ]

        moved_ids = [movement.movement_id for movement, _, _, _, _ in movement_rows]
        reversed_ids = set()
        if moved_ids:
            reversed_ids = set(
                self._execute_bounded_query(
                    session,
                    select(InventoryMovement.reference_id).where(
                        InventoryMovement.movement_type == "reversal",
                        InventoryMovement.reference_id.in_(moved_ids),
                    ),
                    self._MAX_QUERY_ROWS,
                    "equipment history reversals",
                )
            )

        lifecycle_rows: list[list[Any]] = []
        equipment_stats: dict[Any, dict[str, Any]] = {}
        for movement, unit, inventory_item, actor_user_id, actor_name in movement_rows:
            request_context = borrow_context.get(movement.reference_id or "")
            request = request_context[0] if request_context else None
            borrower_name = request_context[1] if request_context else ""
            assignment = assignment_context.get((request.id, unit.id)) if request and unit else None

            lifecycle_rows.append([
                unit.serial_number or "",
                unit.unit_id,
                inventory_item.item_id if inventory_item else "",
                inventory_item.name if inventory_item else "Deleted Inventory Item",
                movement.movement_id,
                self._format_optional_timestamp(movement.occurred_at),
                movement.movement_type,
                movement.qty_change,
                movement.reason_code or "",
                movement.note or "",
                actor_name or actor_user_id or "",
                movement.reference_type or "",
                movement.reference_id or "",
                request.request_id if request else "",
                borrower_name,
                request.customer_name if request else "",
                request.location_name if request else "",
                self._format_optional_timestamp(assignment.released_at if assignment else (request.released_at if request else None)),
                self._format_optional_timestamp(assignment.returned_at if assignment else (request.returned_at if request else None)),
                assignment.condition_on_release if assignment else "",
                assignment.condition_on_return if assignment else "",
                self._format_bool(movement.movement_id in reversed_ids),
            ])

            stats = equipment_stats.setdefault(
                unit.id,
                {
                    "unit": unit,
                    "item": inventory_item,
                    "event_count": 0,
                    "borrow_count": 0,
                    "last_borrowed_at": None,
                    "last_returned_at": None,
                    "last_movement_type": "",
                    "last_movement_at": None,
                },
            )
            stats["event_count"] += 1
            stats["last_movement_type"] = movement.movement_type
            stats["last_movement_at"] = movement.occurred_at
            if movement.movement_type == "borrow_release":
                stats["borrow_count"] += 1
                stats["last_borrowed_at"] = movement.occurred_at
            if movement.movement_type == "borrow_return":
                stats["last_returned_at"] = movement.occurred_at

        equipment_headers = [
            "Item ID",
            "Item Name",
            "Category",
            "Classification",
            "Item Type",
            "Unit ID",
            "Serial Number",
            "Current Status",
            "Current Condition",
            "Expiration Date",
            "Created At",
            "Total Lifecycle Events",
            "Total Borrows",
            "Last Borrowed At",
            "Last Returned At",
            "Last Movement Type",
            "Last Movement At",
            "Deleted",
            "Archived",
        ]
        equipment_rows = [
            [
                stats["item"].item_id if stats["item"] else "",
                stats["item"].name if stats["item"] else "Deleted Inventory Item",
                stats["item"].category if stats["item"] else "",
                stats["item"].classification if stats["item"] else "",
                stats["item"].item_type if stats["item"] else "",
                stats["unit"].unit_id,
                stats["unit"].serial_number or "",
                stats["unit"].status,
                stats["unit"].condition or "",
                stats["unit"].expiration_date.isoformat() if stats["unit"].expiration_date else "",
                self._format_optional_timestamp(stats["unit"].created_at),
                stats["event_count"],
                stats["borrow_count"],
                self._format_optional_timestamp(stats["last_borrowed_at"]),
                self._format_optional_timestamp(stats["last_returned_at"]),
                stats["last_movement_type"],
                self._format_optional_timestamp(stats["last_movement_at"]),
                self._format_bool(stats["unit"].is_deleted),
                self._format_bool(stats["unit"].is_archived),
            ]
            for stats in equipment_stats.values()
        ]

        if format == "csv":
            return self._create_response(event_headers, lifecycle_rows, format, "equipment_history_report")

        movement_type_counts: dict[str, int] = defaultdict(int)
        status_counts: dict[str, int] = defaultdict(int)
        for movement, unit, _, _, _ in movement_rows:
            movement_type_counts[movement.movement_type] += 1
            status_counts[unit.status] += 1
        summary_rows = self._build_filter_summary_rows(
            "Equipment History",
            format,
            timeline_mode,
            anchor_date,
            date_from,
            date_to,
            include_deleted,
            include_archived,
            {
                "Movement Type Filter": movement_type or "",
                "Item Filter": item_id or "",
                "Serial Filter": normalized_serial_number or "",
                "Unique Serials": len(equipment_rows),
                "Lifecycle Events": len(lifecycle_rows),
            },
        )
        for value, count in sorted(movement_type_counts.items()):
            summary_rows.append([f"Movement Type: {value}", count])
        for value, count in sorted(status_counts.items()):
            summary_rows.append([f"Current Status: {value}", count])

        return self._create_multi_sheet_response(
            [
                ("Summary", ["Metric", "Value"], summary_rows),
                ("Equipment", equipment_headers, equipment_rows),
                ("Lifecycle Events", event_headers, lifecycle_rows),
            ],
            "equipment_history_report",
        )

    def export_entrusted(
        self,
        session: Session,
        format: str,
        search: Optional[str] = None,
        status: Optional[str] = None,
        category: Optional[str] = None,
        classification: Optional[str] = None,
    ) -> StreamingResponse:
        from systems.inventory.services.entrusted_item_service import EntrustedItemService
        
        service = EntrustedItemService()
        assignments, _ = service.get_all_entrusted(
            session=session,
            skip=0,
            limit=self._MAX_EXPORT_ROWS,
            search=search,
            status=status,
            category=category,
            classification=classification
        )

        headers = [
            "Assignment ID",
            "Item Name",
            "Serial Number",
            "Assigned To Name",
            "Assigned To ID",
            "Assigned At",
            "Returned At",
            "Status",
            "Notes"
        ]

        data = [
            [
                ass.assignment_id,
                ass.item_name or "N/A",
                ass.serial_number or "",
                ass.assigned_to_name or "Unknown User",
                ass.assigned_to_user_id,
                ass.assigned_at,
                ass.returned_at if ass.returned_at else "N/A",
                "Returned" if ass.returned_at else "Assigned",
                ass.notes or ""
            ]
            for ass in assignments
        ]

        return self._create_response(headers, data, format, "entrusted_items")

    def _create_response(self, headers: List[str], data: List[List[Any]], format: str, prefix: str) -> StreamingResponse:
        filename = f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        sanitized_headers = [self._sanitize_export_cell(header) for header in headers]
        sanitized_data = self._sanitize_export_rows(data)
        
        if format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(sanitized_headers)
            writer.writerows(sanitized_data)
            output.seek(0)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
            )
        
        elif format == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = prefix.replace("_", " ").title()
            
            # Write headers
            for col, header in enumerate(sanitized_headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Write data
            for row_idx, row_data in enumerate(sanitized_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
            )
        
        raise ValueError(
            f"Unsupported export format: {format}. Supported formats: csv, xlsx"
        )
